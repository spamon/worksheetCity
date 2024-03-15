from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.header_footer import HeaderFooter
# import win32api
# import win32print
from openpyxl.utils import get_column_letter



def convert_to_mm(value):
    print(f"Original value: {value}")  # Debug print

    # Try to convert the value based on different measurement types
    match_cm = re.search(r'(\d+(\.\d+)?)\s*cm', value, re.IGNORECASE)
    match_mm = re.search(r'(\d+(\.\d+)?)\s*mm', value, re.IGNORECASE)
    match_inch = re.search(r'(\d+(\.\d+)?)\s*inch', value, re.IGNORECASE)
    
    if match_cm:
        numeric_value = float(match_cm.group(1))
        converted_value = int(numeric_value * 10)  # Convert cm to mm
        print(f"Converted from cm to mm: {converted_value}")  # Debug print
        return converted_value
    elif match_mm:
        numeric_value = float(match_mm.group(1))
        converted_value = int(numeric_value)  # Value is already in mm
        print(f"Value in mm: {converted_value}")  # Debug print
        return converted_value
    elif match_inch:
        numeric_value = float(match_inch.group(1))
        converted_value = int(numeric_value * 25.4)  # Convert inches to mm
        print(f"Converted from inches to mm: {converted_value}")  # Debug print
        return converted_value
    else:
        try:
            numeric_value = float(value)
            print(f"Value in mm: {int(numeric_value)}")  # Debug print
            return int(numeric_value)
        except ValueError:
            print("Conversion error: Invalid numeric value")  # Debug print
            return None




def calculate_sizes_vertical_blinds(product_data):
    # Filter only for Vertical Blinds
    if product_data.get('Product Type') == 'Vertical Blind':
        width_str = product_data.get('Width', '0')
        width = convert_to_mm(width_str)

        louver_drop_size = None

        if width is None:
            product_data['Finished Size'] = 'Invalid Width'
            product_data['Cut Rail Size'] = 'Invalid Width'
            product_data['Qty Louvers'] = ''
            return product_data

        operation_type = product_data.get('Operation Types')
        measurement_type = product_data.get('Measurement Type')

          # Debug print
        print(f"Operation Type: {operation_type}")

        finished_size = cut_rail_size = 0

        if measurement_type == 'Recess':
            finished_size = width - 10
        elif measurement_type == 'Exact':
            finished_size = width
            # Increase louver drop size by 10mm for 'Exact' measurement type
            louver_drop_size = convert_to_mm(product_data.get('Length', '0'))
            if louver_drop_size is not None:
                louver_drop_size += 10

        if operation_type == 'Manual Operation':
            cut_rail_size = finished_size - 20
            product_data['Child Safety'] = 'Side Fix'
        elif operation_type == 'Wand Operation' or operation_type == 'Want Operation':
            print("Processing Want Operation")
            cut_rail_size = finished_size - 10

        product_data['Finished Size'] = f'{finished_size}mm'
        product_data['Cut Rail Size'] = f'{cut_rail_size}mm'
        product_data['Qty Louvers'] = ''
        
        print("Louver Drop Size (Before):", product_data.get('Length'))

        louver_drop_size = convert_to_mm(product_data.get('Length', '0'))
        if louver_drop_size is not None:
            # louver_drop_size += 10  # Example calculation
            product_data['Louver Drop Size'] = f'{louver_drop_size}mm'
        else:
            product_data['Louver Drop Size'] = 'Invalid Length'

        # Debug the louver drop size after calculation
        print("Louver Drop Size (After):", product_data.get('Louver Drop Size'))


    return product_data


def extract_vertical_blind_data(driver, customer_name):
    if "[Sample]" in driver.page_source:
        return None

    table = driver.find_element(By.ID, 'data-table')
    rows = table.find_elements(By.TAG_NAME, 'tr')

    try:
        notes_panel = driver.find_element(By.XPATH, '//div[@class="panel panel-primary"]/div[@class="panel-body"]')
        customer_notes = notes_panel.text.strip()
    except:
        customer_notes = "No notes found."

    all_product_data = []

    for row in rows[1:]:
        product_data = {'Customer Name': customer_name}

        quantity_cell = row.find_element(By.XPATH, './/td[position()=3]')
        quantity = quantity_cell.text.strip()  # Get the quantity as a string

        fabric_name_elements = row.find_elements(By.XPATH, './/td/a')
        fabric_name = fabric_name_elements[1].text.strip() if len(fabric_name_elements) > 1 else 'Unknown Fabric'
        product_data['Fabric Name'] = fabric_name + f" x{quantity}"


        product_details = row.find_elements(By.XPATH, './/div[@class="basket_custom_option"]')
        for detail in product_details:
            label = detail.find_element(By.CLASS_NAME, 'basket_custom_option_label').text.strip(':')
            value = detail.find_element(By.XPATH, './/following-sibling::div').text
            product_data[label] = value
            

        # Call the appropriate calculate_sizes function based on Product Type
        if product_data.get('Product Type') == 'Vertical Blind':
            product_data = calculate_sizes_vertical_blinds(product_data)
        elif product_data.get('Product Type') == 'Roller Blind' and (product_data.get('Roller Type') == 'Standard Roller' or product_data.get('Roller Type') == 'Forward Roll Roller'):
            product_data = calculate_sizes_standard_roller_blinds(product_data)
        elif product_data.get('Product Type') == 'Allusion Blind':
            product_data = calculate_sizes_allusion_blinds(product_data)
        elif product_data.get('Product Type') == 'Roller Blind' and product_data.get('Roller Type') == 'Cassette Roller':
            product_data = calculate_sizes_cassette_roller_blind(product_data)
        elif product_data.get('Product Type') == 'Grip Fit Roller Blind':
            product_data = calculate_sizes_grip_fit_roller_blinds(product_data)
        elif product_data.get('Product Type') == 'Day & Night Blind':
            product_data = calculate_sizes_day_and_night_blinds(product_data)
        elif product_data.get('Product Type') == 'Perfect Fit Pleated':
            product_data = calculate_sizes_perfect_fit_pleated(product_data)
        elif product_data.get('Product Type') == 'Roman Blind':
            product_data = calculate_sizes_roman_blinds(product_data)
        if 'Height required' in product_data:
            product_data = calculate_sizes_vertical_blind_slats(product_data)
        

        all_product_data.append(product_data)

    return all_product_data, customer_notes


def calculate_sizes_allusion_blinds(product_data):
    # Filter only for Allusion Blinds
    if product_data.get('Product Type') == 'Allusion Blind':
        width_str = product_data.get('Width', '0')
        width = convert_to_mm(width_str)

        if width is None:
            product_data['Finished Size'] = 'Invalid Width'
            product_data['Cut Rail Size'] = 'Invalid Width'
            product_data['Qty Louvers'] = ''
            return product_data

        operation_type = product_data.get('Operation Types')
        measurement_type = product_data.get('Measurement Type')

        finished_size = cut_rail_size = 0
        louver_drop_size = None  # Initialize louver_drop_size with None

        if measurement_type == 'Recess':
            finished_size = width - 10
        elif measurement_type == 'Exact':
            finished_size = width
            # Increase louver drop size by 10mm for 'Exact' measurement type
            louver_drop_size = convert_to_mm(product_data.get('Length', '0'))
            if louver_drop_size is not None:
                louver_drop_size += 10

        # For Allusion Blinds, subtract 12mm from the length
        length_str = product_data.get('Length', '0')
        length = convert_to_mm(length_str)
        if length is not None:
            adjusted_length = length - 12
            louver_drop_size = adjusted_length  # Store the adjusted louver drop size in a variable
            product_data['Length'] = f'{adjusted_length}mm'  # Update the Length value
        else:
            louver_drop_size = None
            product_data['Louver Drop Size'] = 'Invalid Length'

        # Check if louver drop size is not None and add it to product_data
        if louver_drop_size is not None:
            product_data['Louver Drop Size'] = f'{louver_drop_size}mm'

        if operation_type == 'Manual Operation':
            cut_rail_size = finished_size - 20
        elif operation_type == 'Want Operation':
            cut_rail_size = finished_size - 10

        product_data['Finished Size'] = f'{finished_size}mm'
        product_data['Cut Rail Size'] = f'{cut_rail_size}mm'
        product_data['Qty Louvers'] = ''
        
        # Check if louver drop size is not None and add it to product_data
        if louver_drop_size is not None:
            product_data['Louver Drop Size'] = f'{louver_drop_size}mm'
        else:
            product_data['Louver Drop Size'] = 'Invalid Length'

        # Update the Length value for Allusion Blinds
        product_data['Length'] = f'{length}mm'

    return product_data




def calculate_sizes_standard_roller_blinds(product_data):
    # Filter for Standard Roller Blinds and Forward Roll Roller Blinds
    if product_data.get('Product Type') == 'Roller Blind' and (product_data.get('Roller Type') == 'Standard Roller' or product_data.get('Roller Type') == 'Forward Roll Roller'):
        width_str = product_data.get('Width', '0')
        width = convert_to_mm(width_str)

        if width is None:
            product_data['Fabric + Rail Width'] = 'Invalid Width'
            product_data['Fabric Drop'] = 'Invalid Length'
            return product_data

        # Calculate Fabric Width and Rail Width
        fabric_width = rail_width = width - 35

        # Calculate Fabric Drop
        length_str = product_data.get('Length', '0')
        length = convert_to_mm(length_str)
        fabric_drop = length + 300

        # Since Fabric Width and Rail Width are identical, use only one of them
        product_data['Fabric + Rail Width'] = f'{fabric_width}mm'

        # Set Fabric Drop
        product_data['Fabric Drop'] = f'{fabric_drop}mm'

        # Remove unwanted keys
        product_data.pop('Fabric Width', None)
        product_data.pop('Rail Width', None)
        product_data.pop('Qty Louvers', None)
        product_data.pop('Measurement Protection', None)

    return product_data




def calculate_sizes_cassette_roller_blind(product_data):
    # Filter only for Cassette Roller Blinds
    if product_data.get('Product Type') == 'Roller Blind' and product_data.get('Roller Type') == 'Cassette Roller':
        width_str = product_data.get('Width', '0')
        width = convert_to_mm(width_str)
        length_str = product_data.get('Length', '0')
        length = convert_to_mm(length_str)

        if width is None or length is None:
            product_data['Cassette Size'] = 'Invalid Width'
            product_data['Tube and Fabric Width'] = 'Invalid Width'
            product_data['Fabric Drop'] = 'Invalid Length'
            return product_data

        measurement_type = product_data.get('Measurement Type')

        if measurement_type == 'Exact':
            cassette_size = width - 4
            tube_and_fabric_width = width - 39
        elif measurement_type == 'Recess':
            cassette_size = width - 14
            tube_and_fabric_width = width - 49

        fabric_drop = length + 300

        product_data['Cassette Size'] = f'{cassette_size}mm'
        product_data['Tube and Fabric Width'] = f'{tube_and_fabric_width}mm'
        product_data['Fabric Drop'] = f'{fabric_drop}mm'

    return product_data


def calculate_sizes_grip_fit_roller_blinds(product_data):
    # Filter only for Grip Fit Roller Blinds
    if product_data.get('Product Type') == 'Grip Fit Roller Blind':
        width_str = product_data.get('Width', '0')
        width = convert_to_mm(width_str)

        if width is None:
            product_data['Cassette Width'] = 'Invalid Width'
            product_data['Tube and Fabric Width'] = 'Invalid Width'
            product_data['Fabric Drop'] = 'Invalid Length'
            return product_data

        # Calculate Cassette Width, Tube and Fabric Width
        cassette_width = width - 20  # Subtract 20mm from the width
        tube_and_fabric_width = width - 42  # Subtract 42mm from the width

        # Calculate Fabric Drop
        length_str = product_data.get('Length', '0')
        length = convert_to_mm(length_str)
        fabric_drop = length + 300  # Add 300mm to the length

        # Add the calculated values to the product_data dictionary
        product_data['Cassette Width'] = f'{cassette_width}mm'
        product_data['Tube and Fabric Width'] = f'{tube_and_fabric_width}mm'
        product_data['Fabric Drop'] = f'{fabric_drop}mm'

        # Remove unwanted keys
        product_data.pop('Qty Louvers', None)
        product_data.pop('Measurement Protection', None)

    return product_data

def calculate_sizes_day_and_night_blinds(product_data):
    # Filter only for Day & Night Blinds
    if product_data.get('Product Type') == 'Day & Night Blind':
        width_str = product_data.get('Width', '0')
        width = convert_to_mm(width_str)

        if width is None:
            product_data['Cassette Width'] = 'Invalid Width'
            product_data['Tube and Fabric Width'] = 'Invalid Width'
            product_data['In Bottom'] = 'Invalid Width'
            product_data['Outer Bottom'] = 'Invalid Width'
            return product_data

        measurement_type = product_data.get('Measurement Type')

        if measurement_type == 'Recess':
            cassette_width = width - 14
            tube_and_fabric_width = cassette_width - 35
            in_bottom = cassette_width - 39
            outer_bottom = cassette_width - 27
        elif measurement_type == 'Exact':
            cassette_width = width - 4
            tube_and_fabric_width = cassette_width - 35
            in_bottom = cassette_width - 39
            outer_bottom = cassette_width - 27

        # The drop is the same as the length
        length_str = product_data.get('Length', '0')
        drop = convert_to_mm(length_str)

        # Update product_data with the calculated values
        product_data['Cassette Width'] = f'{cassette_width}mm'
        product_data['Tube and Fabric Width'] = f'{tube_and_fabric_width}mm'
        product_data['In Bottom'] = f'{in_bottom}mm'
        product_data['Outer Bottom'] = f'{outer_bottom}mm'
        product_data['Drop'] = f'{drop}mm'

        # Remove unwanted keys
        product_data.pop('Measurement Protection', None)

    return product_data

def calculate_sizes_perfect_fit_pleated(product_data):
    # Filter only for Perfect Fit Pleated Blinds
    if product_data.get('Product Type') == 'Perfect Fit Pleated':
        width_str = product_data.get('Width', '0')
        width = convert_to_mm(width_str)
        length_str = product_data.get('Length', '0')
        length = convert_to_mm(length_str)

        if width is None or length is None:
            product_data['Fabric Width'] = 'Invalid Width'
            product_data['Fabric Rail'] = 'Invalid Width'
            product_data['Top and Bottom Frame'] = 'Invalid Width'
            product_data['Side Frame'] = 'Invalid Length'
            product_data['Cells'] = ''
            return product_data

        # Calculate Fabric Width and Fabric Rail
        fabric_width = fabric_rail = width - 16

        # Calculate Top and Bottom Frame
        top_bottom_frame = width - 28

        # Calculate Side Frame
        side_frame = length - 28

        # Add the calculated values to the product_data dictionary
        product_data['Fabric Width'] = f'{fabric_width}mm'
        product_data['Fabric Rail'] = f'{fabric_rail}mm'
        product_data['Top and Bottom Frame'] = f'{top_bottom_frame}mm'
        product_data['Side Frame'] = f'{side_frame}mm'
        product_data['Cells'] = ''  # Placeholder for manual entry

        # Remove unwanted keys
        product_data.pop('Measurement Protection', None)

    return product_data

def calculate_sizes_roman_blinds(product_data):
    if product_data.get('Product Type') == 'Roman Blind':
        width_str = product_data.get('Width', '0')
        width = convert_to_mm(width_str)

        if width is None:
            product_data['Finished Width'] = 'Invalid Width'
            return product_data

        measurement_type = product_data.get('Measurement Type')

        finished_width = width

        if measurement_type == 'Recess':
            finished_width -= 10  # Subtract 10mm for recess measurement

        # Convert drop to mm but don't change the value
        drop_str = product_data.get('Length', '0')
        length = convert_to_mm(drop_str)
        if length is None:
            product_data['Finished Drop'] = 'Invalid Drop'
        else:
            product_data['Finished Length'] = f'{length}mm'

        product_data['Finished Width'] = f'{finished_width}mm'
        
        # Remove unwanted keys
        product_data.pop('Width', None)
        product_data.pop('Length', None)
        product_data.pop('Measurement Protection', None)

    return product_data

def calculate_sizes_vertical_blind_slats(product_data):
    height_str = product_data.get('Height required', '0')
    height = convert_to_mm(height_str)

    if height is None:
        product_data['Adjusted Height'] = 'Invalid Height'
    else:
        adjusted_height = height + 50  # Add 50mm to the height
        product_data['Adjusted Height'] = f'{adjusted_height}mm'

    # # Remove unwanted keys
    # product_data.pop('Height required', None)
    # product_data.pop('Replacement Vertical Slat Weights & Chains required', None)

    return product_data









def main():
    driver = webdriver.Chrome()
    driver.maximize_window()
    wait = WebDriverWait(driver, 10)

    driver.get("https://www.emeraldblindsandcurtains.co.uk/z-admin/login/")
    username_field = driver.find_element(By.CLASS_NAME, "form-control[name='email']")
    password_field = driver.find_element(By.CLASS_NAME, "form-control[name='password']")
    username_field.send_keys("shaun_mcgrath451@btinternet.com")
    password_field.send_keys("")
    password_field.send_keys(Keys.RETURN)

    order_urls = [
        "https://www.emeraldblindsandcurtains.co.uk/z-admin/orders/view/4299/",
        # "https://www.emeraldblindsandcurtains.co.uk/z-admin/orders/view/4211/",
        # "https://www.emeraldblindsandcurtains.co.uk/z-admin/orders/view/4102/",
        # # "https://www.emeraldblindsandcurtains.co.uk/z-admin/orders/view/4132/",
        # "https://www.emeraldblindsandcurtains.co.uk/z-admin/orders/view/4131/"
        # ... (add your other URLs here)
    ]

    for specific_order_url in order_urls:
        driver.get(specific_order_url)
        wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@class="customer-description"]/div[@class="name mb10"]')))
        customer_name = driver.find_element(By.XPATH, '//div[@class="customer-description"]/div[@class="name mb10"]').text.strip()
        order_data, customer_notes = extract_vertical_blind_data(driver, customer_name)

        if order_data is not None:
            grouped_data = {}
            for data in order_data:
                product_type = data.get('Product Type', 'Other')
                if product_type not in grouped_data:
                    grouped_data[product_type] = []
                grouped_data[product_type].append(data)

            safe_customer_name = customer_name.replace(' ', '_', -1).replace('/', '_', -1).replace('\\', '_', -1)
            excel_file_path = f'{safe_customer_name}_extracted_products.xlsx'

            wb = Workbook()
            wb.remove(wb.active)  # Remove the default sheet

            for product_type, data_list in grouped_data.items():
                ws = wb.create_sheet(title=product_type)
                ws.append([product_type])  # Write the product type as a title
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(data_list[0]))  # Merge cells for the title
                ws['A1'].font = Font(bold=True, size=14)  # Set font style for title
                ws['A1'].alignment = Alignment(horizontal='center')

                df = pd.DataFrame(data_list).drop(columns=['Product Type', 'Width', 'Length', 'Measurement Protection', 'Brand'], errors='ignore')

                for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 2):  # Start from row 2
                    ws.append(r)
                    for c_idx, cell in enumerate(r, 1):
                        ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row=r_idx, column=c_idx).border = Border(left=Side(style='thin'),
                                                                        right=Side(style='thin'),
                                                                        top=Side(style='thin'),
                                                                        bottom=Side(style='thin'))
                        if r_idx == 2:
                            ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)

                # Set column widths, skipping the first row
                column_width = 8.5
                for col in range(1, len(data_list[0]) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = column_width

                ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3)
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.fitToWidth = 1

                if customer_notes:
                    ws.append(['Customer Notes:'])
                    ws.append([customer_notes])

            wb.save(excel_file_path)

            
            # printer_name = win32print.GetDefaultPrinter()
            # win32api.ShellExecute(0, "print", excel_file_path, f'/d:"{printer_name}"', ".", 0)

        else:
            print(f"No valid order data extracted from {specific_order_url}.")

    driver.quit()

if __name__ == "__main__":
    main()
